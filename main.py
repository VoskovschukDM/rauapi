import requests
import json
import datetime
import pandas as pd
import numpy as np
import time
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept": "application/json, text/plain, */*",
    "Host": "rau.nalog.gov.ru",
    "Referer": "https://rau.nalog.gov.ru/ngsw-worker.js",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-orign"
}


def get_req_from_file():
    req = {}
    bans = {}
    with open("req.txt") as f:
        s = f.readline()
        while s != '\n':
            arr = s.split()
            if len(arr) != 1:
                req[arr[0]] = (float(arr[1]), float(arr[2]))
            s = f.readline()
        s = f.readline()
        while len(s) != 0:
            arr = s.split()
            if len(arr) != 1:
                bans[arr[0]] = list(map(float, arr[1:]))
            s = f.readline()
    return req, bans


def json_to_df_row(json_data, columns):
    #'guid', 'region', 'dateInitiation', 'shortName', 'fullName', 'ogrn', 'inn', 'kpp', 'okvedCd', 'businessBaseCost', 'businessLiquidationCost', 'repaymentCalc', 'workingCapitalNeed', 'solvencyRnk', 'registrationAddress', 'averageNumber', 'balanceTotal', 'revenue', 'authorizedCapital', 'fixedAssets', 'inventory', 'receivables', 'paymentTaxes'
    result = pd.DataFrame([[
        json_data[columns[0]],
        json_data[columns[1]]['name'],
        json_data[columns[2]],
        json_data['bankrupts'][0][columns[3]],
        json_data['bankrupts'][0][columns[4]],
        json_data['bankrupts'][0][columns[5]],
        json_data['bankrupts'][0][columns[6]],
        json_data['bankrupts'][0][columns[7]],
        json_data['bankrupts'][0]['businessInfo'][columns[8]],
        json_data['bankrupts'][0]['businessInfo'][columns[9]] if columns[9] in json_data['bankrupts'][0]['businessInfo'] else None,
        json_data['bankrupts'][0]['businessInfo'][columns[10]] if columns[10] in json_data['bankrupts'][0]['businessInfo'] else None,
        json_data['bankrupts'][0]['businessInfo'][columns[11]] if columns[11] in json_data['bankrupts'][0]['businessInfo'] else None,
        json_data['bankrupts'][0]['businessInfo'][columns[12]] if columns[12] in json_data['bankrupts'][0]['businessInfo'] else None,
        json_data['bankrupts'][0]['businessInfo'][columns[13]] if columns[13] in json_data['bankrupts'][0]['businessInfo'] else None,
        json_data['bankrupts'][0][columns[14]],
        json_data['bankrupts'][0]['averageQuantity'][columns[15]] if columns[15] in json_data['bankrupts'][0]['averageQuantity'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[16]] if columns[16] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[17]] if columns[17] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[18]] if columns[18] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[19]] if columns[19] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[20]] if columns[20] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[21]] if columns[21] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['bankrupts'][0]['boInfo'][columns[22]] if columns[22] in json_data['bankrupts'][0]['boInfo'] else None,
        json_data['region']['code']
    ]], columns=columns)
    return result



def get_company_info(guid : str):
    url = "https://rau.nalog.gov.ru/backend/api/v2/legal-cases/" + guid
    params = {
        "skip": 0,
        "take": 1,
        "filter.isActiveStatus": "true",
        "dfs": "true"
    }

    r = requests.get(url, params=params, timeout=30)
    return r.json()



def get_list(target_date: datetime.date, columns, requirements, bans):

    result = pd.DataFrame(columns= columns)

    url = "https://rau.nalog.gov.ru/backend/api/v2/legal-cases"
    params = {
        "skip": 0,
        "take": 100,
        "filter.procedureDateFrom": target_date.strftime("%Y-%m-%d"),
        "filter.isActiveStatus": "true",
        "dfs": "true"
    }
    r = requests.get(url, params=params, timeout=30)

    while len(r.json()['items']) != 0:
        for i in r.json()['items']:
            if i['bankrupts'][0]['type'] != 'company':
                continue
            else:
                tmp = get_company_info(i['guid'])
                if len(tmp) != 0:
                    res = json_to_df_row(tmp, columns)
                    if not res.empty:
                        requirements_met = True
                        for j in requirements:
                            if res[j][0] == None:
                                requirements_met = False
                                break
                            if requirements[j][0] != -1 and res[j][0] < requirements[j][0]:
                                requirements_met = False
                                break
                            if requirements[j][1] != -1 and res[j][0] > requirements[j][1]:
                                requirements_met = False
                                break
                        for j in bans:
                            if res[j][0] == None:
                                requirements_met = False
                                break
                            if res[j][0] in bans[j]:
                                requirements_met = False
                                break
                        if requirements_met:
                            result = pd.concat([result, res], ignore_index=True)


        params['skip'] += 100
        r = requests.get(url, params=params, timeout=30)

    return result

retries = 10
columns = ['guid', 'region', 'dateInitiation', 'shortName', 'fullName', 'ogrn', 'inn', 'kpp', 'okvedCd', 'businessBaseCost', 'businessLiquidationCost', 'repaymentCalc', 'workingCapitalNeed', 'solvencyRnk', 'registrationAddress', 'averageNumber', 'balanceTotal', 'revenue', 'authorizedCapital', 'fixedAssets', 'inventory', 'receivables', 'paymentTaxes', 'regionCode']
result = pd.DataFrame(columns= columns)

for i in range(6, -1, -1):
    date = datetime.date.today() - datetime.timedelta(days=i)
    res = pd.DataFrame(columns= columns)

    for attempt in range(retries):
        try:
            requirements, bans = get_req_from_file()
            res = get_list(date, columns, requirements, bans)
            print(f"Success for {date}")
            break
        except Exception as e:
            print(f"Attempt {attempt+1} failed for {date}: {e}")
            time.sleep(2)

    if not res.empty:
        result = pd.concat([result, res], ignore_index=True)

excel_name = f'{(datetime.date.today() - datetime.timedelta(days=6)).strftime("%d-%m-%Y")}to{datetime.date.today().strftime("%d-%m-%Y")}.xlsx'
result.to_excel(excel_name, index=False, engine="openpyxl")

wb = load_workbook(excel_name)
ws = wb.active

width_in_pixels = 100
width_in_chars = width_in_pixels * 0.142857

for col in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = width_in_chars

wb.save(excel_name)